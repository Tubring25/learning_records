import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

def excel_to_html_with_style(excel_file_path, html_output_path):
    # 加载Excel文件，保留样式
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws = wb.active
    
    # 创建HTML表格内容
    html_content = generate_html_table(ws)
    
    # 生成CSS样式
    css_styles = generate_css_styles(ws)
    
    # 组合HTML和CSS
    complete_html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
        {css_styles}
        </style>
    </head>
    <body>
        {html_content}
    </body>
    </html>
    """
    
    # 写入HTML文件
    with open(html_output_path, "w", encoding="utf-8") as html_file:
        html_file.write(complete_html)
    
    print(f"HTML文件已生成: {html_output_path}")

def generate_html_table(ws):
    table_content = "<table class='excel-table'>\n"
    
    # 获取所有合并单元格的范围
    merged_cells = ws.merged_cells.ranges
    
    # 跟踪已处理的合并单元格
    processed_cells = set()
    
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        table_content += "<tr>\n"
        
        for col_idx, cell in enumerate(row, 1):
            # 检查单元格是否在已处理的合并单元格中
            cell_coord = cell.coordinate
            if cell_coord in processed_cells:
                continue
            
            # 获取rowspan和colspan
            rowspan, colspan = 1, 1
            for merged_range in merged_cells:
                if cell_coord in merged_range:
                    rowspan = merged_range.max_row - merged_range.min_row + 1
                    colspan = merged_range.max_col - merged_range.min_col + 1
                    
                    # 将合并范围内的所有单元格标记为已处理
                    for r in range(merged_range.min_row, merged_range.max_row + 1):
                        for c in range(merged_range.min_col, merged_range.max_col + 1):
                            processed_cells.add(f"{get_column_letter(c)}{r}")
                    
                    break
            
            # 生成单元格样式类
            cell_class = f"cell-{row_idx}-{col_idx}"
            
            # 处理单元格内容
            cell_value = cell.value if cell.value is not None else ""
            if isinstance(cell_value, str):
                cell_value = cell_value.replace('\n', '<br>')
            
            # 添加单元格到HTML
            span_attrs = ""
            if rowspan > 1:
                span_attrs += f" rowspan='{rowspan}'"
            if colspan > 1:
                span_attrs += f" colspan='{colspan}'"
            
            table_content += f"    <td class='{cell_class}'{span_attrs}>{cell_value}</td>\n"
        
        table_content += "</tr>\n"
    
    table_content += "</table>"
    return table_content

def generate_css_styles(ws):
    css = """
    .excel-table {
        border-collapse: collapse;
        width: 100%;
    }
    .excel-table td {
        border: 1px solid #ddd;
        padding: 8px;
        vertical-align: top;
    }
    """
    
    # 设置列宽
    for col_idx, col in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        col_width = ws.column_dimensions[col_letter].width
        if col_width:
            css += f"""
    .excel-table td:nth-child({col_idx}) {{
        width: {col_width}px;
    }}
            """
    
    # 设置行高
    for row_idx, row in enumerate(ws.rows, 1):
        row_height = ws.row_dimensions[row_idx].height
        if row_height:
            css += f"""
    .excel-table tr:nth-child({row_idx}) {{
        height: {row_height}px;
    }}
            """
    
    # 设置单元格样式
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for col_idx, cell in enumerate(row, 1):
            cell_class = f".cell-{row_idx}-{col_idx}"
            
            # 初始化样式属性
            style_props = []
            
            # 字体样式
            if cell.font:
                if cell.font.bold:
                    style_props.append("font-weight: bold")
                if cell.font.italic:
                    style_props.append("font-style: italic")
                if cell.font.underline:
                    style_props.append("text-decoration: underline")
                if cell.font.size:
                    style_props.append(f"font-size: {cell.font.size}pt")
                if cell.font.color:
                    rgb = cell.font.color.rgb
                    if rgb:
                        style_props.append(f"color: #{rgb[2:]}")
                if cell.font.name:
                    style_props.append(f"font-family: '{cell.font.name}'")
            
            # 对齐方式
            if cell.alignment:
                if cell.alignment.horizontal:
                    style_props.append(f"text-align: {cell.alignment.horizontal}")
                if cell.alignment.vertical:
                    v_align = cell.alignment.vertical
                    if v_align == 'top':
                        style_props.append("vertical-align: top")
                    elif v_align == 'center':
                        style_props.append("vertical-align: middle")
                    elif v_align == 'bottom':
                        style_props.append("vertical-align: bottom")
            
            # 背景颜色
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                rgb = cell.fill.start_color.rgb
                if rgb != "00000000":  # 不是默认颜色
                    style_props.append(f"background-color: #{rgb[2:]}")
            
            # 边框样式
            if cell.border:
                for side in ['top', 'right', 'bottom', 'left']:
                    border = getattr(cell.border, side)
                    if border and border.style:
                        border_style = "solid" if border.style != "none" else "none"
                        border_color = f"#{border.color.rgb[2:]}" if border.color and border.color.rgb else "#000000"
                        style_props.append(f"border-{side}: 1px {border_style} {border_color}")
            
            # 如果有样式属性，添加到CSS
            if style_props:
                css += f"""
    {cell_class} {{
        {"; ".join(style_props)};
    }}
                """
    
    return css

# 使用示例
if __name__ == "__main__":
    excel_file_path = "input.xlsx"
    html_output_path = "output.html"
    excel_to_html_with_style(excel_file_path, html_output_path)
